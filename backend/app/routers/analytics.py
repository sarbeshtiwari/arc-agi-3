from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
import io

from app.auth import get_admin_user, get_current_user
from app.database import get_db
from app.models import Game, PlaySession, TempGameSession, User
from app.models import now_ist
from app.schemas import DashboardStats, GameAnalyticsResponse, GameResponse
from app.services.analytics import AnalyticsService

router = APIRouter(prefix="/api/analytics", tags=["analytics"])


@router.get("/dashboard", response_model=DashboardStats)
def get_dashboard(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Get dashboard overview statistics."""
    stats = AnalyticsService.get_dashboard_stats(db)
    return DashboardStats(
        total_games=stats["total_games"],
        active_games=stats["active_games"],
        total_plays=stats["total_plays"],
        total_users=stats["total_users"],
        total_wins=stats["total_wins"],
        total_game_overs=stats["total_game_overs"],
        pending_requests=stats["pending_requests"],
        win_rate=stats["win_rate"],
        daily_plays=stats["daily_plays"],
        daily_wins=stats["daily_wins"],
        daily_labels=stats["daily_labels"],
        game_distribution=stats["game_distribution"],
        recent_games=[GameResponse.model_validate(g) for g in stats["recent_games"]],
        top_played_games=stats["top_played_games"],
    )


@router.get("/game/{game_id}")
def get_game_analytics(
    game_id: str,
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Get detailed analytics for a specific game."""
    game = db.query(Game).filter(Game.game_id == game_id).first()
    if not game:
        raise HTTPException(status_code=404, detail="Game not found")

    return AnalyticsService.get_game_analytics(db, game.id, days=days)


@router.get("/sessions")
def get_recent_sessions(
    game_id: str | None = None,
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Get recent play sessions."""
    game_db_id = None
    if game_id:
        game = db.query(Game).filter(Game.game_id == game_id).first()
        if game:
            game_db_id = game.id

    return AnalyticsService.get_recent_sessions(db, game_db_id, limit)


@router.get("/replay/{session_id}")
def get_replay(
    session_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Get replay data for a play session."""
    replay = AnalyticsService.get_replay(db, session_id)
    if not replay:
        raise HTTPException(status_code=404, detail="Session not found")
    return replay


# ──── Temp Game Sessions (admin only) ────
@router.get("/temp-sessions")
def get_temp_sessions(
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """List all temp game sessions (Play Your Own), newest first."""
    sessions = (
        db.query(TempGameSession)
        .order_by(TempGameSession.started_at.desc())
        .all()
    )
    return [
        {
            "id": s.id,
            "session_guid": s.session_guid,
            "game_id": s.game_id,
            "player_name": s.player_name or "Anonymous",
            "state": s.state,
            "score": s.score or 0,
            "total_actions": s.total_actions or 0,
            "current_level": s.current_level or 0,
            "game_overs": s.game_overs or 0,
            "total_time": s.total_time or 0,
            "level_stats": s.level_stats or [],
            "action_log": s.action_log or [],
            "started_at": s.started_at.isoformat() if s.started_at else None,
            "ended_at": s.ended_at.isoformat() if s.ended_at else None,
            "expires_at": s.expires_at.isoformat() if s.expires_at else None,
        }
        for s in sessions
    ]


@router.delete("/temp-sessions")
def delete_all_temp_sessions(
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Delete all temp game sessions."""
    count = db.query(TempGameSession).count()
    db.query(TempGameSession).delete()
    db.commit()
    return {"detail": f"Deleted {count} temp session(s)"}


# ──── Export Sessions as Excel ────
@router.get("/export/{game_id}")
def export_sessions_excel(
    game_id: str,
    filter: str = Query("all", description="Filter: all, today, date, range"),
    date: str = Query(None, description="Date for 'date' filter (YYYY-MM-DD)"),
    date_from: str = Query(None, description="Start date for 'range' filter (YYYY-MM-DD)"),
    date_to: str = Query(None, description="End date for 'range' filter (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Export play sessions for a game as Excel file with date filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    game = db.query(Game).filter(Game.game_id == game_id).first()
    if not game:
        raise HTTPException(status_code=404, detail="Game not found")

    # Build query
    query = db.query(PlaySession).filter(PlaySession.game_id == game.id)

    # Apply date filter
    today = now_ist().replace(hour=0, minute=0, second=0, microsecond=0)

    if filter == "today":
        query = query.filter(PlaySession.started_at >= today)
    elif filter == "date" and date:
        try:
            d = datetime.strptime(date, "%Y-%m-%d")
            query = query.filter(
                PlaySession.started_at >= d,
                PlaySession.started_at < d + timedelta(days=1),
            )
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    elif filter == "range" and date_from and date_to:
        try:
            d_from = datetime.strptime(date_from, "%Y-%m-%d")
            d_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(
                PlaySession.started_at >= d_from,
                PlaySession.started_at < d_to,
            )
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    sessions = query.order_by(PlaySession.started_at.desc()).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{game_id} Sessions"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    # Headers
    headers = [
        "Player", "State", "Level Reached", "Total Actions",
        "Total Time (s)", "Game Overs", "Score",
        "Started At", "Ended At",
        "Level Stats"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, s in enumerate(sessions, 2):
        user = db.query(User).filter(User.id == s.user_id).first() if s.user_id else None
        player = s.player_name or (user.username if user else "Anonymous")

        # Format level stats as readable string
        level_stats_str = ""
        if s.level_stats:
            parts = []
            for ls in s.level_stats:
                status = "Done" if ls.get("completed") else "Incomplete"
                parts.append(
                    f"Lv{ls.get('level', 0) + 1}: {ls.get('actions', 0)}moves, "
                    f"{round(ls.get('time', 0), 1)}s, "
                    f"{ls.get('game_overs', 0)}deaths, "
                    f"{ls.get('resets', 0)}resets ({status})"
                )
            level_stats_str = " | ".join(parts)

        row_data = [
            player,
            s.state or "NOT_FINISHED",
            (s.current_level or 0) + 1,
            s.total_actions or 0,
            round(s.total_time or 0, 2),
            s.game_overs or 0,
            s.score or 0,
            s.started_at.strftime("%Y-%m-%d %H:%M:%S") if s.started_at else "",
            s.ended_at.strftime("%Y-%m-%d %H:%M:%S") if s.ended_at else "",
            level_stats_str,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in (3, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ("Game ID", game_id),
        ("Total Sessions", len(sessions)),
        ("Wins", sum(1 for s in sessions if s.state == "WIN")),
        ("Game Overs", sum(1 for s in sessions if s.state == "GAME_OVER")),
        ("In Progress", sum(1 for s in sessions if s.state == "NOT_FINISHED")),
        ("Total Game Overs (deaths)", sum(s.game_overs or 0 for s in sessions)),
        ("Avg Time (s)", round(sum(s.total_time or 0 for s in sessions) / max(len(sessions), 1), 2)),
        ("Avg Actions", round(sum(s.total_actions or 0 for s in sessions) / max(len(sessions), 1), 1)),
        ("Filter", filter),
        ("Exported At", now_ist().strftime("%Y-%m-%d %H:%M:%S IST")),
    ]
    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell = ws2.cell(row=row_idx, column=1, value=label)
        cell.font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=value)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{game_id}_sessions_{filter}"
    if filter == "date" and date:
        filename += f"_{date}"
    elif filter == "range" and date_from and date_to:
        filename += f"_{date_from}_to_{date_to}"
    filename += ".xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ──── Export ALL Games Sessions as Excel ────
@router.get("/export-all")
def export_all_sessions_excel(
    filter: str = Query("all", description="Filter: all, today, date, range"),
    date: str = Query(None, description="Date for 'date' filter (YYYY-MM-DD)"),
    date_from: str = Query(None, description="Start date for 'range' filter (YYYY-MM-DD)"),
    date_to: str = Query(None, description="End date for 'range' filter (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Export all play sessions across all games as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = db.query(PlaySession)

    today = now_ist().replace(hour=0, minute=0, second=0, microsecond=0)
    if filter == "today":
        query = query.filter(PlaySession.started_at >= today)
    elif filter == "date" and date:
        try:
            d = datetime.strptime(date, "%Y-%m-%d")
            query = query.filter(PlaySession.started_at >= d, PlaySession.started_at < d + timedelta(days=1))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    elif filter == "range" and date_from and date_to:
        try:
            d_from = datetime.strptime(date_from, "%Y-%m-%d")
            d_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(PlaySession.started_at >= d_from, PlaySession.started_at < d_to)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    sessions = query.order_by(PlaySession.started_at.desc()).all()

    # Pre-fetch all games and users for performance
    all_games = {g.id: g for g in db.query(Game).all()}
    all_users = {u.id: u for u in db.query(User).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "All Sessions"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"), right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"), bottom=Side(style="thin", color="E5E7EB"),
    )

    headers = [
        "Game ID", "Player", "State", "Level Reached", "Total Actions",
        "Total Time (s)", "Game Overs", "Score",
        "Started At", "Ended At", "Level Stats",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, s in enumerate(sessions, 2):
        game = all_games.get(s.game_id)
        user = all_users.get(s.user_id) if s.user_id else None
        player = s.player_name or (user.username if user else "Anonymous")
        game_id_str = game.game_id if game else "unknown"

        level_stats_str = ""
        if s.level_stats:
            parts = []
            for ls in s.level_stats:
                status = "Done" if ls.get("completed") else "Incomplete"
                parts.append(f"Lv{ls.get('level', 0)+1}: {ls.get('actions', 0)}m, {round(ls.get('time', 0), 1)}s, {ls.get('game_overs', 0)}deaths, {ls.get('resets', 0)}resets ({status})")
            level_stats_str = " | ".join(parts)

        row_data = [
            game_id_str, player, s.state or "NOT_FINISHED",
            (s.current_level or 0) + 1, s.total_actions or 0,
            round(s.total_time or 0, 2), s.game_overs or 0, s.score or 0,
            s.started_at.strftime("%Y-%m-%d %H:%M:%S") if s.started_at else "",
            s.ended_at.strftime("%Y-%m-%d %H:%M:%S") if s.ended_at else "",
            level_stats_str,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in (4, 5, 6, 7, 8):
                cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    games_with_sessions = {}
    for s in sessions:
        g = all_games.get(s.game_id)
        gid = g.game_id if g else "unknown"
        if gid not in games_with_sessions:
            games_with_sessions[gid] = {"plays": 0, "wins": 0, "game_overs_total": 0}
        games_with_sessions[gid]["plays"] += 1
        if s.state == "WIN":
            games_with_sessions[gid]["wins"] += 1
        games_with_sessions[gid]["game_overs_total"] += (s.game_overs or 0)

    ws2.cell(row=1, column=1, value="Game ID").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Sessions").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Wins").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="Total Deaths").font = Font(bold=True)
    for i, (gid, stats) in enumerate(games_with_sessions.items(), 2):
        ws2.cell(row=i, column=1, value=gid)
        ws2.cell(row=i, column=2, value=stats["plays"])
        ws2.cell(row=i, column=3, value=stats["wins"])
        ws2.cell(row=i, column=4, value=stats["game_overs_total"])

    row_offset = len(games_with_sessions) + 3
    summary_data = [
        ("Total Sessions", len(sessions)),
        ("Total Games", len(games_with_sessions)),
        ("Filter", filter),
        ("Exported At", now_ist().strftime("%Y-%m-%d %H:%M:%S IST")),
    ]
    for i, (label, value) in enumerate(summary_data):
        ws2.cell(row=row_offset + i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row_offset + i, column=2, value=value)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"all_games_sessions_{filter}"
    if filter == "date" and date:
        filename += f"_{date}"
    elif filter == "range" and date_from and date_to:
        filename += f"_{date_from}_to_{date_to}"
    filename += ".xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ──── Export Games List as Excel ────
@router.get("/export-games")
def export_games_list_excel(
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Export all games with their details as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    games = db.query(Game).order_by(Game.created_at.desc()).all()
    all_users = {u.id: u for u in db.query(User).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Games"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"), right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"), bottom=Side(style="thin", color="E5E7EB"),
    )

    headers = [
        "Game ID", "Name", "Status", "Version", "Game Code",
        "Description", "Game Rules", "Owner",
        "Drive Link", "Video Link",
        "Default FPS", "Tags", "Levels",
        "Total Plays", "Total Wins", "Avg Score",
        "Uploaded By", "Created At",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, g in enumerate(games, 2):
        uploader = all_users.get(g.uploaded_by) if g.uploaded_by else None
        num_levels = len(g.baseline_actions) if g.baseline_actions else 0
        tags_str = ", ".join(g.tags) if g.tags else ""

        row_data = [
            g.game_id, g.name,
            "Active" if g.is_active else "Inactive",
            g.version, g.game_code,
            g.description or "", g.game_rules or "", g.game_owner_name or "",
            g.game_drive_link or "", g.game_video_link or "",
            g.default_fps or 5, tags_str, num_levels,
            g.total_plays or 0, g.total_wins or 0, round(g.avg_score or 0, 2),
            uploader.username if uploader else "",
            g.created_at.strftime("%Y-%m-%d %H:%M:%S") if g.created_at else "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in (11, 13, 14, 15, 16):
                cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"games_list_{now_ist().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
